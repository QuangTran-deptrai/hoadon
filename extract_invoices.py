
import os
import re
import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Category mapping based on extracted services
CATEGORY_KEYWORDS = {
    "Ăn uống": [
        "món", "lẩu", "gà", "bò", "heo", "cá", "cua", "mực", "tôm", "ghẹ", "sò", 
        "gỏi", "xào", "nướng", "chiên", "hấp", "hầm", "quay", "cơm", "xôi", "soup",
        "trà", "cà phê", "nước", "coca", "matcha", "oolong", "trái cây", "bánh",
        "đậu", "trứng", "lươn", "hàu", "khô mực", "khăn lạnh", "hủ tiếu", "baba",
        "bồ câu", "chả", "dừa", "khoáng", "suối", "sả","Rượu"
    ],
    "Viễn thông": [
        "cước", "di động", "thẻ cào", "sim", "điện thoại", "internet", "mạng", "mệnh giá","THE CAO MENH GIA"
    ],
    "Dịch vụ IT": [
        "cài đặt", "máy tính", "sửa chữa", "bảo trì", "setup", "install", "văn phòng"
    ],
    "Thuê phòng": [
        "thuê phòng", "phòng số", "cho thuê phòng", "phòng họp", "meeting room"
    ],
    "Vận chuyển": [
        "thuê xe", "vận chuyển", "taxi", "grab", "giao hàng", "shipping", "ôtô"
    ],
    "Hoa/Quà tặng": [
        "hoa tươi", "hoa", "quà", "gift"
    ],
    "Phụ tùng/Thiết bị": [
        "tay đẩy", "thiết bị", "phụ tùng", "linh kiện"
    ]
}

def classify_content(services_text):
    """Classify services into categories using keyword matching."""
    if not services_text:
        return "Khác"
    text_lower = services_text.lower()
    
    scores = {}
    for category, keywords in CATEGORY_KEYWORDS.items():
        score = sum(1 for kw in keywords if kw.lower() in text_lower)
        if score > 0:
            scores[category] = score
    
    if scores:
        return max(scores, key=scores.get)
    return "Khác"


def format_price_value(value):
    
    if not value or not isinstance(value, str):
        return value
    
    # Remove any spaces
    value = value.strip()
    
    # Check if it has Vietnamese decimal format (,00 or ,0 at end)
    if ',' in value:
        # Split by comma - left part is integer, right part is decimal
        parts = value.rsplit(',', 1)
        integer_part = parts[0]
        # Ignore decimal part
    else:
        integer_part = value
    
    # Convert dots to nothing (remove thousands separator), then parse as integer
    integer_part = integer_part.replace('.', '')
    
    try:
        num = int(integer_part)
        # Format with comma as thousands separator
        return f"{num:,}"
    except ValueError:
        return value  # Return original if can't parse

COMMON_UNITS = {
    "CÁI", "CHIẾC", "BỘ", "GÓI", "HỘP", "THÙNG", "BAO", "CHAI", "LON", "LÍT", "LIT", "KG", "GRAM", "GM", "MÉT", 
    "M", "M2", "M3", "CUỘN", "TẤM", "THANH", "VIÊN", "VỈ", "TỜ", "QUYỂN", "CUỐN", "RAM", "CẶP", "ĐÔI", 
    "DĨA", "ĐĨA", "PHẦN", "THỐ", "TÔ", "CHÉN", "LY", "CỐC", "SUẤT", "KIM", "CHẬU", "CÂY", "GIỜ", "NGÀY", "THÁNG", 
    "NĂM", "LẦN", "CHUYẾN", "LƯỢT", "PHÚT", "KW", "KWH", "SỐ", "MÓN", "KỆ", "BỊCH", "NỒI", "CON", "PCS", "NGƯỜI"
}

# Keywords indicating item lines to skip (headers, footers, summaries)
JUNK_TEXT_KEYWORDS = [
    'stt', 'tên hàng', 'đơn vị tính', 'số lượng', 'thành tiền', 'người mua', 
    'ký bởi', 'trang', 'thuế suất', 'cộng tiền', 'tổng cộng', 'bằng chữ',
    'tiền thuế', 'serial', 'ký hiệu', 'mẫu số', 'vnd', 'chuyển khoản',
    'vat invoice', 'đơn vị bán', 'mã tra cứu', 'vat) rate)', 'vat rate', 
    'gtgt', 'rate)', 'amount)', 'rate%)', 'tên h', 'đơ n v', 's ố l', 
    'vị tính', 'sau thuế', 'chiết khấu', 'a b c'
]

# Keywords for detecting surcharge/fee items
SURCHARGE_KEYWORDS = ['phụ thu', 'phí dịch vụ', 'phí phục vụ', 'service charge', 'surcharge']


def is_junk_text(text):
    """Check if text is a header/footer/summary line that should be skipped."""
    if not text or len(text) < 2:
        return True
    t = text.lower()
    if re.match(r'^([A-Z]\s+)+[A-Z]$', text):
        return True
    if re.match(r'^[\d\s()=x+]+$', text):
        return True
    if any(w in t for w in JUNK_TEXT_KEYWORDS):
        return True
    if len(t) > 50 and sum(1 for c in t if c in '()') > 4:
        return True
    return False


def parse_money(value):
    """Parse Vietnamese money string to integer. Returns None if invalid."""
    if not value:
        return None
    try:
        clean = str(value).replace('.', '').replace(',', '')
        return int(clean)
    except (ValueError, TypeError):
        return None


def format_money(value):
    """Format integer as money string with comma separator."""
    if value is None:
        return ""
    return f"{value:,}"


def parse_vietnamese_number(value):
    """Parse Vietnamese number format (dot as thousand separator, comma as decimal)."""
    if not value:
        return 0
    try:
        return float(str(value).replace('.', '').replace(',', '.'))
    except (ValueError, TypeError):
        return 0


def extract_services_from_text(full_text):
    """Extract service/product details with qty, unit_price, and amount."""
    services = []
    lines = full_text.split('\n')

    for line_idx, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue
            
        first_token = line.split()[0] if line.split() else ""
        if not first_token.isdigit():
            continue
            
        # Ignore column headers "1 2 3 4 5"
        if re.match(r'^[\d\s.,|()x=+]+$', line): continue
        if re.match(r'^[A-Z\s]+[\d\s=x]+$', line): continue
             
        # Must start with number (1-3 digits) - STT
        match_start = re.match(r'^(\d{1,3})[._\-\s|]+', line)
        if not match_start: continue
        
        # Find all number blocks
        num_pattern = r'(\d+(?:[.,]\d+)*)'
        all_nums = list(re.finditer(num_pattern, line))
        
        if len(all_nums) < 3:  # Need at least STT + qty + amount
            # Exception: surcharge items may have only STT + amount (2 numbers)
            line_lower = line.lower()
            is_surcharge_line = any(kw in line_lower for kw in SURCHARGE_KEYWORDS)
            if not (is_surcharge_line and len(all_nums) == 2):
                continue
        
        stt_end = len(match_start.group(0))
        
        # Strategy: Find the LAST UNIT word (from COMMON_UNITS) that is followed by numbers
        # This handles cases where unit words appear in description (like "từ ngày 15/12")
        
        tokens = line.split()
        
        line_after_stt = line[stt_end:]
        unit_idx = -1  # index of unit word in tokens
        
        # Find LAST unit word that has numbers after it
        for i in range(len(tokens) - 1, -1, -1):  # iterate backwards
            tok = tokens[i]
            # Ignore THANH as unit (usually start of name like Thanh long)
            if tok.upper() in COMMON_UNITS and tok.upper() != 'THANH':
                # Check if there are numbers after this position
                remaining = ' '.join(tokens[i+1:])
                if re.search(r'\d+(?:[.,]\d+)*', remaining):
                    unit_idx = i
                    break
        
        if unit_idx == -1:
            # No unit found - find name boundary by looking for first "price-like" number
            # A price typically has 4+ digits or contains decimal separator
            nums_after_stt = [m for m in all_nums if m.start() > stt_end]
            
            # Find first number that looks like a price (>= 1000 or has comma/dot)
            name_end_pos = len(line)
            for m in nums_after_stt:
                num_str = m.group(1)
                # Check if it's a price number: contains . or , separator OR is >= 4 digits
                if '.' in num_str or ',' in num_str or len(num_str.replace('.', '').replace(',', '')) >= 4:
                    name_end_pos = m.start()
                    break
            
            if name_end_pos <= stt_end:
                continue
            name_part = line[stt_end:name_end_pos].strip()
            nums = [m.group(1) for m in nums_after_stt]
        else:
            # Unit found - take name as tokens before unit, numbers after unit
            name_tokens = tokens[:unit_idx]
            
            # Special case: if the last token before unit is a 4-digit year (e.g., "2025"), include it in name
            # This handles cases like "tháng 11 năm 2025 Tháng 1 1.260.000"
            if unit_idx > 0 and re.match(r'^\d{4}$', tokens[unit_idx - 1]):
                # Keep the year in name, effectively making unit_idx point to one position after
                pass  # name_tokens already includes the year since it's tokens[:unit_idx]
            
            # Instead of offset calculation which is error-prone, just take tokens after unit
            remaining_tokens = tokens[unit_idx+1:]
            remaining_text = " ".join(remaining_tokens)
            
            nums_iter = re.finditer(r'(\d+(?:[.,]\d+)*)', remaining_text)
            nums = [m.group(1) for m in nums_iter]
            
            if len(nums) < 2:
                # Check if it's a surcharge item before skipping
                temp_name = ' '.join(name_tokens).lower()
                is_surcharge = any(kw in temp_name for kw in SURCHARGE_KEYWORDS)
                if not (is_surcharge and len(nums) == 1):
                    continue
                    
            name_part = ' '.join(name_tokens)
        
        # Special handling for surcharge/fee items with only one number (just amount, no qty/price)
        # e.g., "19 Phụ thu 171.500"
        is_surcharge_item = any(kw in name_part.lower() for kw in SURCHARGE_KEYWORDS)
        
        if len(nums) < 2:
            if is_surcharge_item and len(nums) == 1:
                # Surcharge with single amount - use it as both price and amount, qty = 1
                nums = ['1', nums[0], nums[0]]  # [qty, unit_price, amount]
            else:
                continue
        
        # Clean name - remove unit words from start and end
        tokens = name_part.split()
        # Remove from end - only exact match units
        while tokens and tokens[-1].upper() in COMMON_UNITS:
            tokens = tokens[:-1]
        # Remove from start (rare but can happen with multi-line merge) - only exact match
        # EXCEPTION: "THANH" is a unit but also start of "Thanh long", "Thanh toán"... so don't remove it at start
        while tokens and tokens[0].upper() in COMMON_UNITS and tokens[0].upper() != "THANH":
            tokens = tokens[1:]
        # Also clean partial unit patterns like "Phần）" or "）Nấm"  
        if tokens:
            # Clean first token if it's just punctuation + text
            first = tokens[0]
            if first.startswith('）') or first.startswith(')'):
                tokens[0] = first[1:].strip()
            # Clean last token if it ends with LONG unit (3+ chars) to avoid cutting "Nấm" -> "Nấ"
            last = tokens[-1] if tokens else ''
            for unit in COMMON_UNITS:
                if len(unit) >= 3 and last.upper().endswith(unit) and len(last) > len(unit):
                    tokens[-1] = last[:-len(unit)].rstrip('（(')
                    break
            # Clean tokens that are JUST unit+bracket like "Phần）" or "Phần)"
            # Only remove if it's strictly unit+bracket, to avoid removing "Nửa phần）"
            # Clean tokens that are JUST unit+bracket like "Phần）" or "Phần)" removal logic removed 
            # as it truncated valid name parts like "Bắp Mỹ（Nửa phần）" where "phần）" is part of the name.
            pass
        name_part = " ".join(tokens)
        
        # Remove STT from the start of name_part if detected
        # This prevents cases like "4 Phần" becoming "Đậu phụ... 4 Phần" after merge
        stt_val = match_start.group(1)
        if name_part.startswith(stt_val):
             name_part = name_part[len(stt_val):].strip()
        
        # Multi-line description handling: merge PREVIOUS and NEXT lines if they look like continuations
        
        # Check PREVIOUS lines for description prefix - but only if current name looks incomplete
        # Conditions for needing prev line: starts lowercase, starts with '(' or ')', or is very short
        # Ignore leading STT number for this check
        check_name = re.sub(r'^\d+\s+', '', name_part).strip()
        first_char = check_name[0] if check_name else ''
        needs_prev = (
            len(check_name) < 5 or  # Very short name
            (first_char and first_char.islower()) or  # Starts lowercase = continuation
            (first_char == '(' or first_char == ')') or  # Starts with paren
            (')' in check_name[:10])  # Has closing paren early
        )
        
        if needs_prev:
            prev_parts = []
            for offset in range(1, 3):  # Check up to 2 lines back
                if line_idx - offset >= 0:
                    prev_line = lines[line_idx - offset].strip()
                    
                    # Stop if: empty, STT line, junk, or has too many numbers
                    if not prev_line or len(prev_line) < 2:
                        break
                    match_stt = re.match(r'^(\d{1,3})[._\-\s|]+', prev_line)
                    if match_stt:
                         # If STT line contains ONLY STT and text (no other numbers), merge it
                         stt_len = len(match_stt.group(0))
                         rest_of_line = prev_line[stt_len:].strip()
                         
                         # Check if remaining part has numbers (likely prices) -> then it's a separate item -> Break
                         if re.search(r'\d', rest_of_line):
                             break
                             
                         # If valid text, use it (strip STT)
                         prev_line = rest_of_line
                    if is_junk_text(prev_line):
                        break
                        
                    # Check nums, but allow dates (DD/MM/YYYY or DD-MM-YYYY)
                    # Remove dates from line before counting nums
                    temp_line = re.sub(r'\d{1,2}[/-]\d{1,2}[/-]\d{2,4}', '', prev_line)
                    
                    # Only count "price-like" numbers (with comma/period separators) as stop conditions
                    # This allows alphanumeric codes like "63A 17235" (vehicle plates) to pass through
                    price_nums = re.findall(r'\d{1,3}(?:[.,]\d{3})+(?:[.,]\d{2})?', temp_line)
                    simple_nums = re.findall(r'\b\d{4,}\b', temp_line)  # Long standalone numbers (4+ digits)
                    
                    if price_nums or len(simple_nums) > 1:
                        break
                    if any(x in prev_line.lower() for x in ['cộng tiền', 'tổng cộng', 'thuế', 'thành tiền']):
                        break
                    
                    # STOP if prev_line looks like the TAIL of a previous item
                    if (prev_line.endswith(')') or prev_line.endswith('）')) and re.search(r'[a-zA-Z]', prev_line):
                        if '(' not in prev_line and '（' not in prev_line: 
                            break
                    
                    # STOP merging if we hit an English paren line that precedes a Vietnamese line
                    # Case Item 7: "Thêm cơm trắng" (collected) <-- "(Sichuan tofu)" (checking). Stop.
                    # Case Item 5: "(Braised..." (collected) <-- "Chân nấm..." (checking). Continue.
                    if prev_parts:
                        last_collected = prev_parts[0] # This is the line conceptually "below" the current prev_line
                        is_vietnamese_line = last_collected[0].isupper() and not last_collected.startswith('(')
                        is_english_paren = prev_line.startswith('(') and re.search(r'[a-zA-Z]', prev_line)
                        if is_vietnamese_line and is_english_paren:
                            break

                    prev_parts.insert(0, prev_line)  # Insert at beginning to maintain order
                else:
                    break
            
            if prev_parts:
                name_part = " ".join(prev_parts) + " " + name_part
        
        # Check NEXT lines for description suffix
        last_char = name_part[-1] if name_part else ''
        
        # Detect if next line is a parenthetical suffix
        next_line_is_suffix = False
        if line_idx >= 0 and line_idx + 1 < len(lines):
            peek_line = lines[line_idx + 1].strip()
            if peek_line.startswith('(') and re.search(r'[a-zA-Z]', peek_line):
                next_line_is_suffix = True
            elif peek_line.startswith('(') and re.search(r'(ngày|từ|đến|tháng|năm)', peek_line, re.IGNORECASE):
                next_line_is_suffix = True

        has_unclosed_paren = (name_part.count('(') > name_part.count(')')) or (name_part.count('（') > name_part.count('）'))
        
        needs_next = (
            last_char in '(-（' or
            name_part.rstrip().endswith('(') or
            name_part.rstrip().endswith('（') or
            has_unclosed_paren or
            next_line_is_suffix
        )
        
        if needs_next:
            next_parts = []
            for offset in range(1, 4):  # Check up to 3 lines ahead
                if line_idx >= 0 and line_idx + offset < len(lines):
                    next_line = lines[line_idx + offset].strip()
                    
                    # Stop if: empty, STT line, junk, or has too many numbers
                    if not next_line or len(next_line) < 2:
                        break
                    if re.match(r'^\d{1,3}[._\-\s|]+', next_line):
                        break
                    if is_junk_text(next_line):
                        break
                    # Remove date patterns before counting numbers
                    temp_next = re.sub(r'\d{1,2}\s*[/-]\s*\d{1,2}\s*[/-]\s*\d{2,4}', '', next_line)
                    nums_in_next = re.findall(r'\d+(?:[.,]\d+)*', temp_next)
                    if len(nums_in_next) > 1:
                        break
                    if any(x in next_line.lower() for x in ['cộng tiền', 'tổng cộng', 'thuế', 'thành tiền']):
                        break
                    
                    # STOP if next_line starts a new item (Vietnamese text, not closing paren)
                    if next_line[0].isupper() and next_line[0] != '(':
                        matches_closing_paren = has_unclosed_paren and ('）' in next_line or ')' in next_line)
                        if not matches_closing_paren:
                            if re.search(r'[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩòóọỏõôồốộổỗơờớợởỡùúụủũưừứựửữỳýỵỷỹđ]', next_line, re.IGNORECASE):
                                break
                    
                    next_parts.append(next_line)
                else:
                    break
            
            if next_parts:
                name_part = name_part + " " + " ".join(next_parts)
        
        name_part = name_part.strip()
        
        # Clean up junk prefixes and suffixes
        prefixes_to_remove = ["GTGT", "VAT) rate)", "VAT rate", "Rate)", "A B C", "khấu", "KHẤU", 
                             "Phần）", "Phần)", "PHẦN）", "PHẦN)", "ĐVT:", "ĐVT"]
        for prefix in prefixes_to_remove:
            if name_part.startswith(prefix + " "):
                name_part = name_part[len(prefix):].strip()
            if name_part.startswith(prefix):
                name_part = name_part[len(prefix):].strip()
        name_part = re.sub(r'^(?:[A-C]\s)+[\d\s=x]+', '', name_part)
        name_part = re.sub(r'^[\d\s=x+]+(\s|$)', '', name_part)
        
        # Remove trailing junk like "- 6 + 7 9 10 = 8 x 9" or "Nồi 2 400.000"
        # Pattern: ends with numbers/operators/units that look like column data
        name_part = re.sub(r'\s+[\-\+]\s*[\d\s=x\+\-]+$', '', name_part)
        # Remove trailing numbers ONLY if they're not:
        # 1. A 4-digit year (like 2025)  
        # 2. Preceded by price/value keywords (like "mệnh giá 20.000")
        trailing_match = re.search(r'(\S+)\s+(\d+[\s.,\d]*)$', name_part)
        if trailing_match:
            word_before = trailing_match.group(1).lower()
            num_part = trailing_match.group(2).strip().split()[0] if trailing_match.group(2).strip() else ""
            keep_number = False
            # Keep 4-digit years
            if len(num_part) == 4 and num_part.isdigit():
                keep_number = True
            # Keep if preceded by price/value keywords
            price_keywords = ['giá', 'gia', 'mệnh', 'số', 'phòng', 'room', 'no', 'no.']
            if word_before in price_keywords:
                keep_number = True
            if not keep_number:
                name_part = name_part[:trailing_match.start(2)].strip()
        
        name_part = name_part.strip()
        
        # FINAL cleanup: remove unit words that appeared after merge
        final_tokens = name_part.split()
        while final_tokens and final_tokens[-1].upper() in COMMON_UNITS:
            final_tokens = final_tokens[:-1]
        # EXCEPTION: "THANH" at start
        while final_tokens and final_tokens[0].upper() in COMMON_UNITS and final_tokens[0].upper() != "THANH":
            final_tokens = final_tokens[1:]
        name_part = " ".join(final_tokens)
        
        # Skip if name is junk
        if len(name_part) < 3:
            continue
        if is_junk_text(name_part):
            continue
        if re.match(r'^[\d\s]+$', name_part):
            continue
        if re.match(r'^[\d\s=x\+\-\.,()\[\]]+$', name_part):
            continue
        
        # Determine qty, unit_price, amount
        # Handle different formats:
        # - Standard (4-5 nums): Qty, Price, Amount, [Rate%, VAT] - Thành tiền at position 3 (index 2)
        # - With discount (6+ nums): Qty, Price, Discount(0), Amount, Rate%, VAT, Total - if nums[2]=0, use nums[3]
        qty = ""
        unit_price = ""
        amount = ""
        
        if len(nums) >= 5:
            # Check if there's a discount column (nums[2] is 0 or 0,00)
            discount_val = nums[2].replace('.', '').replace(',', '').strip('0')
            has_discount_column = (discount_val == '' or discount_val == '0')
            
            qty = nums[0]
            unit_price = nums[1]
            
            # Smart Selection: Compare candidates nums[2] and nums[3] against Expected = Qty * Price
            # This handles cases like: Qty, Price, ServiceCharge, Amount (HĐ 63518)
            cand2 = nums[2]
            cand3 = nums[3] if len(nums) >= 4 else None
            
            try:
                q_val = float(qty.replace(',', '.').replace('.', '') if ',' in qty else qty)
                p_val = float(unit_price.replace('.', '').replace(',', '.'))
                if p_val < 100 and '.' in unit_price:
                    p_val = float(unit_price.replace('.', '').replace(',', ''))
                
                expected = q_val * p_val
                if expected == 0 and q_val == 0:
                    expected = p_val
                
                v2 = parse_vietnamese_number(cand2)
                v3 = parse_vietnamese_number(cand3) if cand3 else 0
                
                diff2 = abs(v2 - expected)
                diff3 = abs(v3 - expected) if cand3 else float('inf')
                
                # Select amount candidate closest to expected value
                if cand3 and diff3 < diff2 and v2 < (0.5 * expected):
                    amount = cand3
                elif has_discount_column and cand3:
                    amount = cand3
                else:
                    amount = cand2
            except:
                # Fallback to logic based on discount column
                if has_discount_column and len(nums) >= 4:
                    amount = nums[3]
                else:
                    amount = nums[2]
        elif len(nums) >= 3:
            qty = nums[0]
            unit_price = nums[1]
            amount = nums[2]
        elif len(nums) == 2:
            qty = nums[0]
            amount = nums[1]
            unit_price = amount
        else:
             continue
        
        # Heuristic fix for "Phí dịch vụ" case where Amount is misidentified as Tax Rate (e.g. 8)
        # If extracted Amount is very small (<= 100) and Unit Price is substantial (> 1000), swap/fix it.
        try:
            # Simple parse assuming checking for small integers vs large price
            a_str = amount.replace('.', '').replace(',', '')
            p_str = unit_price.replace('.', '').replace(',', '')
            if a_str.isdigit() and p_str.isdigit():
                 val_amount = float(a_str)
                 val_price = float(p_str)
                 # Check raw values (accounting for potential decimal scaling issues, but 8 vs 46800 is clear)
                 if val_amount <= 100 and val_price > 1000:
                      amount = unit_price
                      if qty == '0' or not qty:
                           qty = "1"
        except:
            pass
        
        services.append({
            "name": name_part,
            "qty": format_price_value(qty),
            "unit_price": format_price_value(unit_price),
            "amount": format_price_value(amount)
        })

    return services

def extract_invoice_data(pdf_source, filename=None):
    """
    Extract invoice data from a PDF file source.
    :param pdf_source: File path (str) or file-like object (BytesIO)
    :param filename: Original filename (if pdf_source is a stream)
    """
    if isinstance(pdf_source, str):
        filename = os.path.basename(pdf_source)
    elif filename is None:
        filename = "Unknown.pdf"

    data = {
        "Tên file": filename,
        "Ngày hóa đơn": "",
        "Số hóa đơn": "",
        "Đơn vị bán": "",
        "Phân loại": "",
        "Số tiền trước Thuế": "",
        "Tiền thuế": "",
        "Số tiền sau": "",
        "Link lấy hóa đơn": "",
        "Mã tra cứu": "",
        "Mã số thuế": "",
        "Mã CQT": "",
        "Ký hiệu": ""
    }
    # Store line items separately for multi-row expansion
    line_items = []
    
    try:
        # Read text directly from PDF using pdfplumber
        full_text = ""
        with pdfplumber.open(pdf_source) as pdf:
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    full_text += page_text + "\n"
        
        # Fallback only works if we have a local file path
        if not full_text and isinstance(pdf_source, str):
            print(f"  Empty PDF text, checking for fallback text file...")
            base_name = os.path.splitext(os.path.basename(pdf_source))[0]
            folder = os.path.dirname(pdf_source)
            # Find closest matching text file (e.g. filename_00001.txt)
            for f in os.listdir(folder):
                # Check for files starting with the base name (ignoring the (1) vs (1)_0001 differences sometimes)
                # Simple check: startswith base_name and ends with .txt
                if f.startswith(base_name) and f.lower().endswith('.txt') and not f.startswith('debug_'):
                    txt_path = os.path.join(folder, f)
                    print(f"  -> Found fallback text file: {f}")
                    try:
                        # Try UTF-8 first
                        with open(txt_path, 'r', encoding='utf-8') as tf:
                            full_text = tf.read()
                        break
                    except UnicodeDecodeError:
                        try:
                             # Try CP1252 / ANSI
                             with open(txt_path, 'r', encoding='cp1252') as tf:
                                full_text = tf.read()
                             break
                        except Exception as e:
                             print(f"  -> Error reading fallback file (encoding): {e}")
                    except Exception as e:
                        print(f"  -> Error reading fallback file: {e}")

        if not full_text.strip():
            print(f"  Could not extract text (scanned PDF?): {filename}")
            # Set all fields to "không nhận diện được"
            for key in data:
                if key != "Tên file":
                    data[key] = "không nhận diện được"
            return data, []  # Return empty line_items

        
        # Extract services from text
        services = extract_services_from_text(full_text)
        
        # ============ EXTRACT FIELDS WITH MULTIPLE PATTERNS ============
        
        # DATE - Multiple patterns
        date_patterns = [
            r'Ngày\s*\(date\)\s*(\d+)\s*tháng\s*\(month\)\s*(\d+)\s*năm\s*\(year\)\s*(\d+)',
            r'Ngày\s*\(day\)\s*(\d+)\s*tháng\s*\(month\)\s*(\d+)\s*năm\s*\(year\)\s*(\d+)',
            r'Ngày\s*\(Date\)\s*(\d+)\s*[Tt]háng\s*\([Mm]onth\)\s*(\d+)\s*[Nn]ăm\s*\([Yy]ear\)\s*(\d+)',
            r'Ngày\s+(\d+)\s+tháng\s+(\d+)\s+năm\s+(\d+)',  # Sapo: Ngày 14 tháng 8 năm 2025
            r'Ngày\s*(\d+)\s*tháng\s*(\d+)\s*năm\s*(\d+)',
            r'Ngày(\d+)tháng(\d+)năm(\d+)',  # No spaces
        ]
        for pattern in date_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                day, month, year = match.groups()
                data["Ngày hóa đơn"] = f"{day}/{month}/{year}"
                break
        
        # INVOICE NUMBER - Multiple patterns (order matters - more specific first)
        inv_patterns = [
            r'Số\s*\(No\.?\)[:\s]*(\d{5,})',  # M-INVOICE: Số(No.): 00007155 (at least 5 digits)
            r'Số[/\s]*\(Invoice No\.?\)[:\s]*(\d+)',
            r'\(RESTAURANT BILL\)\s*(\d+)',  # VNPT Restaurant: (RESTAURANT BILL) 00004501
            r'Số:\s*(\d+)',  # Explicit colon: Số: 00007155
            r'Số hóa đơn[:\s]*(\d+)',
            r'Số\s*\(No\.?\)[:\s]*(\d+)',  # M-INVOICE with any digits
            r's[éèẹẽe][: ]+\s*(\d+)',  # OCR typo: sé (Petrolimex)
            r'S[óố][: ]+\s*(\d+)',  # OCR typo: Só/Số
            # NOTE: Removed generic 'Số[:\s]+(\d+)' - too broad, matches addresses
        ]
        for pattern in inv_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                data["Số hóa đơn"] = match.group(1)
                break
        
        # Fallback: Extract from filename if missing
        if not data["Số hóa đơn"]:
            # Try to find a long number in filename?
            fname = os.path.splitext(os.path.basename(pdf_path))[0]
            # Split by underscores or hyphens
            parts = re.split(r'[_\-\s]', fname)
            # Filter for pure digit sequences, reasonable length (e.g. >3)
            # Avoid parts that look like dates if possible, but simplest is last long number
            nums = [p for p in parts if p.isdigit() and len(p) > 2]
            if nums:
                 data["Số hóa đơn"] = nums[-1] # Take the last number found pattern often has invoice num at end

        seller_patterns = [
            r'Đơn vị bán hàng\s*\([Ss]eller\)[:\s]*(.+)',  # M-INVOICE format
            r'Tên người bán\s*\([Ss]eller\)[:\s]*(.+)',  # VNPT format
            r'Đơn vị bán hàng\s*\([Cc]ompany\)[:\s]*(.+)',  # MISA variation
            r'Đơn vị bán hàng[:\s]*(.+)',  # Simple format (Petrolimex)
            r'Đơn vị bán\s*\([Ss]eller\)[:\s]*(.+)',
            r'Tên đơn vị bán hàng[:\s]*(.+)',
            r'HỘ KINH DOANH[:\s]*(.+)',
            # NOTE: Removed 'Người bán' pattern - it captures 'Người bán hàng(Seller)' incorrectly
        ]
        for pattern in seller_patterns:
            match = re.search(pattern, full_text)
            if match:
                seller = match.group(1).strip()
                
                # Check for multi-line split (common in VNPT)
                # e.g. "CHI NHÁNH... (LOẠI HÌNH DOANH NGHIỆP:\nCÔNG TY TNHH)..."
                # Find start and end index of this match in full_text
                start_idx = match.end(1)
                # Look ahead for next line
                rest_of_text = full_text[match.end():]
                next_line_match = re.match(r'\n([^\n]+)', rest_of_text)
                if next_line_match:
                    next_line = next_line_match.group(1).strip()
                    # Heuristic to merge:
                    # 1. Seller line ends with ':', '(', or "DOANH NGHIỆP"
                    # 2. Next line starts with "CÔNG TY", "TẬP ĐOÀN", ")"
                    if (seller.endswith(':') or seller.endswith('(') or 'DOANH NGHIỆP' in seller[-15:]):
                         seller = seller + " " + next_line
                    
                # Clean up - remove (Seller): prefix and other text
                # Normalize newlines to spaces just in case
                seller = seller.replace('\n', ' ')
                
                # Robust cleanup of "Seller" / "Company" prefixes
                # Removes: "(Seller):", "Seller :", "(Company):", "Doanh nghiệp:", etc.
                seller = re.sub(r'^\s*[\(\[]?\s*(?:Seller|Company|Người bán|Doanh nghiệp|Tên đơn vị)\s*[\)\]]?\s*[:\.\-]?\s*', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'^\s*[:\.\-]+\s*', '', seller) # Clean remaining colons/dashes
                seller = re.sub(r'\s*Mã số thuế.*$', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'\s*MST.*$', '', seller, flags=re.IGNORECASE)
                seller = re.sub(r'\s*Địa chỉ.*$', '', seller, flags=re.IGNORECASE)
                
                # Check for invalid seller content (captured footer text/codes)
                # Added 'địa chỉ', 'address' to prevent grabbing Address line
                if any(x in seller.lower() for x in ['mã nhận hóa đơn', 'code for checking', 'tra cứu tại', 'địa chỉ', 'address']):
                    continue
                
                # Check for placeholder capture
                if seller.lower().replace(':', '').strip() in ['(seller)', 'seller', 'người bán', 'tên đơn vị']:
                    continue
                    
                if len(seller) > 5:
                    data["Đơn vị bán"] = seller
                    break
        
        # PRIORITY FALLBACK 1: First line(s) before first "Mã số thuế" - this is most reliable for MISA invoices
        # where seller company name is at the very top of the document
        if not data["Đơn vị bán"]:
            # Find the position of first "Mã số thuế"
            mst_pos = full_text.find("Mã số thuế")
            if mst_pos > 0:
                # Get text before first MST
                text_before_mst = full_text[:mst_pos].strip()
                lines_before_mst = [l.strip() for l in text_before_mst.split('\n') if l.strip()]
                
                # First non-empty line that looks like a company name
                for line in lines_before_mst[:6]:  # Check first 6 lines to handle headers
                    # Must contain company keywords AND be reasonably long
                    if len(line) > 10 and any(kw in line.upper() for kw in ['CÔNG TY', 'TẬP ĐOÀN', 'CHI NHÁNH', 'NHÀ HÀNG', 'DNTN', 'HỘ KINH DOANH']):
                        # Exclude headers and BUYER info
                        if not any(bad in line.upper() for bad in ['HÓA ĐƠN', 'CỘNG HÒA', 'ĐỘC LẬP', 'TÊN NGƯỜI MUA', 'TÊN ĐƠN VỊ:', 'PHÂN PHỐI TỔNG HỢP DẦU KHÍ']):
                            # Multi-line company name: if next line is also uppercase text, merge
                            idx = lines_before_mst.index(line)
                            if idx + 1 < len(lines_before_mst):
                                next_line = lines_before_mst[idx + 1]
                                # Merge if next line doesn't contain MST markers and is short uppercase
                                if next_line and 'Mã số' not in next_line and 'Địa chỉ' not in next_line:
                                    if (next_line.isupper() or (len(next_line) < 40 and ':' not in next_line)) and 'PHÂN PHỐI' not in next_line.upper():
                                        line = line + " " + next_line
                            data["Đơn vị bán"] = line
                            break

        # FALLBACK 2: Ký bởi (Signed by) - common in footer, company name may span multiple lines
        if not data["Đơn vị bán"]:
            # Try multi-line pattern first: "Ký bởi:CÔNG TY...\nTHẾ THÊM"
            sign_match = re.search(r'(?:Ký bởi|Được ký bởi)[:\s]*([A-ZĐ][A-ZĐÀÁẢÃẠ\s]+(?:\n[A-ZĐÀÁẢÃẠ\s]+)?)', full_text)
            if sign_match:
                signer = sign_match.group(1).replace('\n', ' ').strip()
                # Only accept if it looks like a company name
                if len(signer) > 5 and any(x in signer.upper() for x in ['CÔNG TY', 'TẬP ĐOÀN', 'CHI NHÁNH', 'NHÀ HÀNG', 'DNTN']):
                    if not any(x in signer.lower() for x in ['địa chỉ', 'address', 'mã số']):
                        data["Đơn vị bán"] = signer
        
        # SERIAL NUMBER (Ký hiệu) - Multiple patterns INCLUDING "Series"
        serial_patterns = [
            r'[KK]ý hiệu\s*/\s*\([Ss]erial(?:\s*No\.?)?\)[:\s]*([A-Z0-9]+)',  # Format with slash: Ký hiệu/ (Serial No)
            r'[KK]ý hiệu\s*\([Ss]erial\)[:\s]*([A-Z0-9]+)',  # VNPT: Ký hiệu(Serial): 1K25THA
            r'[KK]ý hiệu\s*\([Ss]erial(?:\s*No\.?)?\)[:\s]*([A-Z0-9]+)',  # M-INVOICE
            r'[KK]ý hiệu\s*\([Ss]eries\)[:\s]*([A-Z0-9]+)',  # VNPT uses "Series"
            r'[KK]ý hiệu[:\s]*([A-Z0-9]+)',
            r'Mẫu số\s*-\s*[KK]ý hiệu[^:]*[:\s]*([A-Z0-9]+)',
        ]
        for pattern in serial_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                data["Ký hiệu"] = match.group(1)
                break
        
        # SECURITY CODE (Mã tra cứu) - Multiple patterns
        security_patterns = [
            r'Mã nhận hóa đơn\s*\([Cc]ode for checking\)[:\s]*([A-Z0-9]+)',  # Special case
            r'Mã tra cứu\s*\([Ll]ookup\s*code\)[:\s]*([A-Za-z0-9_]+)',  # VNPT: Mã tra cứu(Lookup code):HCM...
            r'Mã tra cứu hóa đơn\s*\([Ii]nvoice code\)[:\s]*([A-Za-z0-9_]+)',  # MISA variation
            r'Mã tra cứu(?:\s*HĐĐT)?(?:\s*này)?[:\s]*([A-Za-z0-9_]+)',
            r'Mã tra cứu\(Invoice code\)[:\s]*([A-Za-z0-9_]+)',  # MISA no-space
            r'Mã số bí mật[:\s]*([A-Za-z0-9_]+)',
            r'Security Code\)[:\s]*([A-Z0-9]+)',
            r'Mã tra cứu[:\s]*([A-Za-z0-9]+)',
            r'[Ll]ookup\s*code[):\s]*([A-Za-z0-9]+)',  # Standalone
        ]
        for pattern in security_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                code = match.group(1)
                # Allow longer codes (VNPT uses 32 chars)
                if 5 <= len(code) <= 35:
                    data["Mã tra cứu"] = code
                    break
                elif len(code) > 35:
                    # Very long code might be CQT, store separately
                    if not data["Mã CQT"]:
                        data["Mã CQT"] = code
        
        # TAX CODE (MST đơn vị bán) - Look for seller's tax code (first one)
        tax_patterns = [
            r'Mã số thuế\s*\([Tt]ax\s*code\)[:\s]*([\d\-\u00AD\s]+)',  # Added \s for spaced numbers
            r'(?:MST|Mã số thuế)[/\s]*\([Tt]ax [Cc]ode\)[:\s]*([\d\-\u00AD\s]+)',
            r'MST/CCCD[^:]*[:\s]*([\d\-\u00AD\s]+)',
            r'(?:MST|Mã số thuế)[:\s]*([\d\-\u00AD\s]+)',
        ]
        tax_codes = []
        for pattern in tax_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            # Clean up matches - remove soft hyphens AND spaces
            cleaned_matches = []
            for m in matches:
                clean = m.replace('\u00AD', '').replace(' ', '').strip()
                # Verify it looks like a tax code (at least 10 chars, digits/hyphens)
                if len(clean) >= 10 and any(c.isdigit() for c in clean):
                    cleaned_matches.append(clean)
            tax_codes.extend(cleaned_matches)
        
        if len(tax_codes) >= 1:
            data["Mã số thuế"] = tax_codes[0]  # Seller's tax code (first one)
        
        # CQT CODE - Multiple patterns (include soft hyphen \u00AD used in some PDFs)
        cqt_patterns = [
            r'Mã\s*(?:của\s*)?[Cc]ơ quan thuế[:\s]*([A-Za-z0-9\-\u00AD]+)',  # M-INVOICE
            r'Mã\s*(?:của\s*)?[Cc]ơ quan thuế\s*\([Tt]ax authority code\)[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'Mã\s*CQT\s*\([Cc]ode\)[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'Mã\s*CQT[:\s]*([A-Za-z0-9\-\u00AD]+)',
            r'Tax authority code[:\s]*([A-Za-z0-9\-\u00AD]+)',
        ]
        for pattern in cqt_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                # Replace soft hyphen with regular hyphen
                cqt_code = match.group(1).strip().replace('\u00AD', '-')
                data["Mã CQT"] = cqt_code
                break
        
        # LOOKUP LINK - Multiple patterns
        link_patterns = [
            r'Tra cứu hóa đơn tại\s*\([^)]+\)[:\s]*(https?://[^\s]+)',  # VNPT: Tra cứu hóa đơn tại (Lookup the invoice at):https://...
            r'Tra cứu hóa đơn tại[:\s]*(https?://[^\s]+)',  # Simple format
            r'(?:Tra cứu[^:]*tại|Trang tra cứu|website)[:\s]*(https?://[^\s]+)',
            r'(https?://[^\s]*(?:tracuu|tra-cuu|invoice|vnpt-invoice|minvoice)[^\s]*)',
        ]
        for pattern in link_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                link = match.group(1).rstrip('.').rstrip(',')
                data["Link lấy hóa đơn"] = link
                break
        
        # AMOUNTS - Multiple patterns
        # Before tax
        before_tax_patterns = [
            r'Cộng tiền hàng[^:]*[:\s]*([\d\.,]+)',
            r'Cộng ti[êề]n hàng[^:]*[:\s]*([\d\.,]+)', # OCR typo: tiên
            r'Tổng tiền chưa thuế[^:]*[:\s]*([\d\.,]+)',  # M-INVOICE
            r'Thành ti[êềẫ]n trước thuế[^:]*[:\s]*([\d\.,]+)', # OCR typo: tiễn
            r'Amount before VAT[^:]*[:\s]*([\d\.,]+)',
            r'[Tt]otal amount[^:]*[:\s]*([\d\.,]+)',
            r'Sub total[^:]*[:\s]*([\d\.,]+)',
        ]
        for pattern in before_tax_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            if matches:
                # Take the LAST match as it's likely the grand total on the last page
                data["Số tiền trước Thuế"] = matches[-1]
                break
        
        # VAT AMOUNT (Tiền thuế)
        vat_patterns = [
            r'Tổng tiền thuế GTGT \d+%[:\s]*([\d\.,]+)', # Specific rate line
            r'\|?Tiền thu[êế] GTGT\s*\(\s*\d+\s*%\s*\)\s*([\d\.,]+)', # MOST SPECIFIC: |Tiền thuê GTGT ( 8% ) 59.265
            r'\|?Tiền thu[êế] GTGT[^:]*[:\s]+(\d[\d\.,]+)', # |Tiền thuê GTGT: 59.265
            r'Tiền thuế\s*\(VAT\s*Amount\)[^:]*[:\s]*([\d\.,]+)',
            r'Tổng tiền thuế[^:]*[:\s]*([\d\.,]+)',  # M-INVOICE
            r'Tiền thu[êế] GTGT[^:]*[:\s]+(\d[\d\.,]+)', # OCR typo: thuê (ensure starts with digit)
            r'VAT amount[^:]*[:\s]*([\d\.,]+)',
            r'Cộng tiền thuế GTGT[^:]*[:\s]*([\d\.,]+)',
        ]
        for pattern in vat_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            if matches:
                 # Take the LAST match
                data["Tiền thuế"] = matches[-1]
                break
        
        # After tax (total payment)
        after_tax_patterns = [
            r'Tổng tiền chịu thuế suất.*[:\s]*[\d\.,]*%\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # M-INVOICE table summary (Total amount): 8% 655.000 52.400 707.400
            r'Tổng cộng\s*\(Total amount\)\s*[:]\s*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # File 2025_812...
            r'Tổng\s*cộng\s*\([Tt]otal\)?[:\s]*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # MISA: Tổng cộng(Total): 375.000 30.000 405.000
            r'Tổngcộng[:\s]*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # No-space: Tổngcộng: 2.816.100 256.158 3.072.258
            r'Tổng cộng\s*[:]\s*([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)', # SAPO/EasyInvoice: Tổng cộng: [Before] [VAT] [Total]
            r'Tổng cộng tiền thanh toán\s*\(Total amount\)\s*([\d\.,\s]+)', # Golden Gate: 5 numbers, take the last one
            r'[Tt]ổng\s*tiền\s*thanh\s*toán\s*\([^)]+\)[:\s]*([\d\.,]+)',  # MISA: Tổng tiền thanh toán (Total amount): 1.800.000
            r'[IT].{1,3}ng\s*số\s*ti[êề]n\s*thanh\s*toán[:\s]*([\d\.,]+)', # OCR typo: Iông, tiên (Petrolimex) - Flexible match for Tổng/Iông
            r'Cộng tiền hàng hóa, dịch vụ[:\s]*[\d\.,]+\s+[\d\.,]+\s+([\d\.,]+)', # File 1226-TK-200k (Total on next line)
            r'[Tt]ổng\s*cộng\s*tiền\s*thanh\s*toán[^:]*[:\s]*([\d\.,]+)',
            r'[Tt]otal\s*payment[^:]*[:\s]*([\d\.,]+)',
            r'TỔNG CỘNG TIỀN THANH TOÁN[^:]*[:\s]*([\d\.,]+)',
            r'Tổng cộng[:\s]+([\d\.,]+)\s+[\d\.,]+\s+([\d\.,]+)',  # Multi-page format
            r'thuế suất:\s*\d+%\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)',  # Tax rate line
        ]
        for pattern in after_tax_patterns:
            matches = list(re.finditer(pattern, full_text, re.IGNORECASE))
            if matches:
                match = matches[-1] # Take the LAST match
                
                # Handle multi-column format (before_tax, vat, after_tax)
                if match.lastindex and match.lastindex >= 3:
                    # Specific check for SAPO/EasyInvoice where Group 1=Before, Group 2=VAT, Group 3=Total
                    # Overwrite existing values as summary line is more reliable
                    data["Số tiền sau"] = match.group(match.lastindex)
                    data["Tiền thuế"] = match.group(2)
                    data["Số tiền trước Thuế"] = match.group(1)
                elif match.lastindex and match.lastindex >= 2:
                    data["Số tiền sau"] = match.group(match.lastindex)
                    # For format with 2 columns, be careful about overwriting
                    if not data["Số tiền trước Thuế"]:
                        data["Số tiền trước Thuế"] = match.group(1)
                else:
                    # Single group or Golden Gate complex case
                    val = match.group(1)
                    # Use helper split if it looks like multiple numbers (Golden Gate)
                    parts = val.strip().split()
                    if len(parts) >= 3 and all(c in '0123456789.,' for c in ''.join(parts)):
                         # Golden Gate: Total amount 1.656.000 100.000 1.556.000 124.480 1.680.480
                         data["Số tiền sau"] = parts[-1]
                         data["Tiền thuế"] = parts[-2]
                         data["Số tiền trước Thuế"] = parts[-3]
                    else:
                         data["Số tiền sau"] = val
                break
        
        # SPECIAL CASE: Hộ Kinh Doanh with Tax Reduction Note (Nghị quyết 204/2025/QH15)
        # e.g. "Cộng tiền bán hàng hóa, dịch vụ: 2.289.962" -> This is the final amount to pay
        if not data["Số tiền sau"] or not data["Số tiền trước Thuế"]:
             # Check for "Cộng tiền bán hàng hóa, dịch vụ" which is common in direct sales invoices
             direct_sales_match = re.search(r'Cộng tiền bán hàng hóa, dịch vụ[:\s]*([\d\.,]+)', full_text, re.IGNORECASE)
             if direct_sales_match:
                 amount = direct_sales_match.group(1)
                 # If we haven't set "Số tiền sau", use this. 
                 # Usually for Hộ Kinh Doanh, total payment = total goods amount (minus discount if any, but usually final)
                 if not data["Số tiền sau"]:
                     data["Số tiền sau"] = amount
                 if not data["Số tiền trước Thuế"]:
                     data["Số tiền trước Thuế"] = amount
                 # If extracted "Tiền thuế" is empty, it might be 0 or calculated from reduction note, 
                 # but usually direct sales don't list VAT separately like deductive invoices. 
                 # We leave VAT empty or 0 if not found.

        # SPECIAL PATTERN: Before Tax + VAT on one line (File 1226-TK-200k.pdf)
        # Cộng tiền hàng hóa, dịch vụ: 219.907 17.593
        # Prioritize this summary line as it matches User's preferred values (rounded)
        double_match = re.search(r'Cộng tiền hàng hóa, dịch vụ[:\s]*([\d\.,]+)\s+([\d\.,]+)', full_text, re.IGNORECASE)
        if double_match:
             # Check if the second number looks like money (digits/dots)
             # Force overwrite to ensure we get the summary values
             data["Số tiền trước Thuế"] = double_match.group(1)
             data["Tiền thuế"] = double_match.group(2)
        
        # For SALES INVOICE (no VAT): if Số tiền sau is empty but we have before tax amount
        if not data["Số tiền sau"] and data["Số tiền trước Thuế"]:
            if "SALES INVOICE" in full_text or "HÓA ĐƠN BÁN HÀNG" in full_text:
                data["Số tiền sau"] = data["Số tiền trước Thuế"]
            else:
                sales_match = re.search(r'Cộng tiền bán hàng[^:]*[:\s]*([\d\.,]+)', full_text)
                if sales_match:
                    data["Số tiền sau"] = sales_match.group(1)
        
        # Calculate and validate money values
        # Validate money values - must be >= 1000
        for col in ["Số tiền trước Thuế", "Tiền thuế", "Số tiền sau"]:
            val = parse_money(data[col])
            if val is not None and val < 1000:
                data[col] = ""  # Invalid, clear it
        
        # Calculate Số tiền sau if not found but we have before tax and VAT
        if not data["Số tiền sau"]:
            before = parse_money(data["Số tiền trước Thuế"])
            vat = parse_money(data["Tiền thuế"])
            if before is not None and vat is not None:
                data["Số tiền sau"] = format_money(before + vat)
            elif before is not None and vat is None:
                # No VAT, total = before tax
                data["Số tiền sau"] = data["Số tiền trước Thuế"]
        
        # REVERSE CASE: If we have Số tiền sau (total) but no Số tiền trước Thuế (before tax)
        # and no VAT was found, then this is a non-VAT invoice, so set pre-tax = post-tax
        if data["Số tiền sau"] and not data["Số tiền trước Thuế"]:
            after = parse_money(data["Số tiền sau"])
            vat = parse_money(data["Tiền thuế"])
            if after is not None and (vat is None or vat == 0):
                # No VAT found or VAT is 0, so pre-tax = post-tax
                data["Số tiền trước Thuế"] = data["Số tiền sau"]
            elif after is not None and vat is not None:
                # VAT exists, so calculate pre-tax = post-tax - VAT
                data["Số tiền trước Thuế"] = format_money(after - vat)
        
        # Store line items for multi-row expansion
        if services:
            line_items = services  # services is now list of dicts with name, qty, unit_price, amount
            
    except Exception as e:
        print(f"Error processing {filename}: {e}")
    
    return data, line_items


def format_excel_output(file_path):
    """Format the Excel output file with professional styles and merge cells."""
    print(f"Applying professional formatting to {file_path}...")
    try:
        wb = openpyxl.load_workbook(file_path)
        if "Hóa đơn" in wb.sheetnames:
            ws = wb["Hóa đơn"]
        else:
            ws = wb.active

        # Define Styles
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
        header_fill = PatternFill("solid", fgColor="4F81BD")  # Professional Blue
        border_style = Side(style='thin', color="000000")
        border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
        
        # Column widths for simplified layout:
        # A=Tên file, B=Ngày, C=Số HĐ, D=Đơn vị bán, E=Phân loại
        # F=Trước thuế, G=Thuế, H=Sau thuế, I=Link
        # J=Mã tra cứu, K=MST, L=Mã CQT, M=Ký hiệu
        widths = {
            'A': 30, 'B': 12, 'C': 15, 'D': 40, 'E': 18,
            'F': 18, 'G': 15, 'H': 18, 'I': 15,
            'J': 20, 'K': 15, 'L': 15, 'M': 12
        }
        
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width

        # Format Header Row (row 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add Filter
        ws.auto_filter.ref = ws.dimensions
        
        # No need for merge logic since each invoice is now one row
        
        # Format Data Rows - borders and number formatting
        money_cols_idx = [6, 7, 8]  # F, G, H (Trước thuế, Thuế, Sau thuế)
        center_cols_idx = [2, 3, 5, 11, 13]  # B, C, E, K, M
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                # Skip merged cells (they don't have col_idx)
                if not hasattr(cell, 'col_idx'):
                    continue
                    
                cell.border = border
                cell.font = Font(name="Arial", size=10)
                
                # Number format for money columns
                if cell.col_idx in money_cols_idx:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif cell.col_idx in center_cols_idx:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        wb.save(file_path)
        print(" -> Formatting complete with merged cells.")
        
    except Exception as e:
        print(f"Error formatting Excel: {e}")
        import traceback
        traceback.print_exc()

def main():
    # Fix Windows console encoding for Vietnamese characters
    import sys
    sys.stdout.reconfigure(encoding='utf-8')
    
    # Input folder for PDF files - users should place new invoices here
    input_folder = r"D:\hoadon\invoices_input"
    
    # Output folder for exported Excel
    output_folder = r"D:\hoadon"
    
    # Create input folder if not exists
    if not os.path.exists(input_folder):
        os.makedirs(input_folder)
        print(f"Created input folder: {input_folder}")
        print("Please add PDF invoice files to this folder and run again.")
        return
    
    # Get all PDF files from input folder
    pdf_files = [f for f in os.listdir(input_folder) if f.lower().endswith('.pdf')]
    
    if not pdf_files:
        print(f"No PDF files found in: {input_folder}")
        print("Please add PDF invoice files to this folder and run again.")
        return
        
    print(f"Processing {len(pdf_files)} PDF files from: {input_folder}\\n")
    
    all_rows = []  # Will contain expanded rows (one per line item)
    
    for pdf_file in pdf_files:
        pdf_path = os.path.join(input_folder, pdf_file)
        print(f"Processing: {pdf_file}")
        
        data, line_items = extract_invoice_data(pdf_path)
        
        # Classify invoice based on line items
        if line_items:
            all_item_names = " ".join([item.get("name", "") for item in line_items])
            data["Phân loại"] = classify_content(all_item_names)
        else:
            data["Phân loại"] = "Khác"
        
        all_rows.append(data)
        
        # Show status
        item_count = len(line_items) if line_items else 0
        seller_display = data['Đơn vị bán'][:30] if data['Đơn vị bán'] and len(data['Đơn vị bán']) > 0 else 'N/A'
        print(f"  -> Ngay: {data['Ngày hóa đơn']}, So: {data['Số hóa đơn']}, Category: {data['Phân loại']}, DonViBan: {seller_display}...")
    
    # Create DataFrame
    df = pd.DataFrame(all_rows)
    
    # Reorder columns
    columns = [
        "Tên file", "Ngày hóa đơn", "Số hóa đơn", "Đơn vị bán", "Phân loại",
        "Số tiền trước Thuế", "Tiền thuế", "Số tiền sau", "Link lấy hóa đơn",
        "Mã tra cứu", "Mã số thuế", "Mã CQT", "Ký hiệu"
    ]
    df = df[columns]
    
    # Format money columns - convert to number and add comma separators
    money_columns = ["Số tiền trước Thuế", "Tiền thuế", "Số tiền sau"]
    for col in money_columns:
        def convert_to_number(x):
            if pd.isna(x) or x == '':
                return None
            x_str = str(x).replace('.', '').replace(',', '')
            try:
                return int(float(x_str))
            except (ValueError, TypeError):
                return x
        df[col] = df[col].apply(convert_to_number)
    
    # Export to Excel
    output_file = os.path.join(output_folder, "hoadon_tonghop.xlsx")
    try:
        df.to_excel(output_file, index=False, sheet_name="Hóa đơn")
    except PermissionError:
        print(f"\nWARNING: Could not save to '{output_file}' because it is open.")
        output_file = os.path.join(output_folder, "hoadon_tonghop_new.xlsx")
        print(f"Saving to '{output_file}' instead.")
        df.to_excel(output_file, index=False, sheet_name="Hóa đơn")
    
    # Print summary
    print(f"\n{'='*50}")
    print(f"SUMMARY:")
    unique_files = df["Tên file"].nunique()
    print(f"  Total rows: {len(df)} (from {unique_files} invoices)")
    for col in ["Ngày hóa đơn", "Số hóa đơn", "Đơn vị bán", "Phân loại", "Số tiền sau"]:
        empty_count = (df[col] == '').sum() + df[col].isna().sum()
        pct = (1 - empty_count/len(df)) * 100
        print(f"  {col}: {pct:.0f}% filled ({len(df)-empty_count}/{len(df)})")
    print(f"\nExported to: {output_file}")
    
    # Apply professional formatting
    format_excel_output(output_file)

if __name__ == "__main__":
    main()
