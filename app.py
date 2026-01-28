import streamlit as st
import pandas as pd
import io
import os
import logging
import sys
import re
from extract_invoices import extract_invoice_data, classify_content
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Configure logging to stdout
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(message)s',
    handlers=[logging.StreamHandler(sys.stdout)]
)
logger = logging.getLogger(__name__)

# Configure page
st.set_page_config(page_title="Invoice Extractor", page_icon="🧾", layout="wide")

# Category options for dropdown
CATEGORY_OPTIONS = [
    "Tự động nhận diện",  # Auto-detect based on invoice content
    "Dịch vụ ăn uống",
    "Dịch vụ phòng nghỉ", 
    "Hoa tươi",
    "Thẻ cào điện thoại",
    "Xăng xe",
    "Quà tặng",
    "Khác (Nhập tay)"
]

# Initialize Session State
if "processing_complete" not in st.session_state:
    st.session_state["processing_complete"] = False
if "processed_df" not in st.session_state:
    st.session_state["processed_df"] = None

# --- Main Application Logic (no login required) ---

# Sidebar
with st.sidebar:
    st.markdown("**Invoice Extractor**")
    st.markdown("---")
    st.caption("Phan tich va trich xuat du lieu tu hoa don PDF")

# App Title
st.title("Invoice Extraction Tool")

# --- WIZARD FLOW ---

if st.session_state["processing_complete"] and st.session_state["processed_df"] is not None:
        # === STEP 4: RESULTS & EXPORT ===
        st.markdown("### ✅ Kết quả xử lý")
        
        col_res1, col_res2 = st.columns([1, 4])
        with col_res1:
            if st.button("⬅️ Làm việc với file khác"):
                st.session_state["processing_complete"] = False
                st.session_state["processed_df"] = None
                st.rerun()
        
        df = st.session_state["processed_df"]
        
        # Excel Export with Merge Logic
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Hóa đơn")
            worksheet = writer.sheets["Hóa đơn"]
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
            header_fill = PatternFill("solid", fgColor="4F81BD")
            border_style = Side(style='thin', color="000000")
            border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
            
            # Column widths
            widths = {'A': 15, 'B': 15, 'C': 12, 'D': 15, 'E': 15, 'F': 30, 'G': 18, 
                      'H': 15, 'I': 12, 'J': 10, 'K': 15, 'L': 18, 'M': 35}
            for col_letter, width in widths.items():
                worksheet.column_dimensions[col_letter].width = width

            # Format Header
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Format Data
            money_cols_idx = [8, 9, 11]  # H, I, K
            center_cols_idx = [1, 2, 3, 4, 5, 10]  # A-E, J
            
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    if isinstance(cell, openpyxl.cell.cell.MergedCell): continue
                    cell.border = border
                    cell.font = Font(name="Arial", size=10)
                    if cell.col_idx in money_cols_idx:
                        cell.number_format = '#,##0'
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    elif cell.col_idx in center_cols_idx:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(vertical="center", wrap_text=True)
            
            # Merge cells for multi-tax-rate invoices
            # Columns to merge by filename: H(Trước VAT), I(VAT), K(Sau thuế)
            # Team column (A) ALWAYS merged by Team value
            merge_by_file_cols = [8, 9, 11]  # H, I, K
            
            # First: Merge Team column by Team value (column A = 1)
            if len(df) > 0:
                start_row = 2
                current_team = worksheet.cell(row=2, column=1).value
                
                for excel_row in range(3, worksheet.max_row + 2):
                    if excel_row > worksheet.max_row:
                        cell_value = None
                    else:
                        cell_value = worksheet.cell(row=excel_row, column=1).value
                    
                    if cell_value != current_team:
                        end_row = excel_row - 1
                        if end_row > start_row:
                            worksheet.merge_cells(f"A{start_row}:A{end_row}")
                            top_cell = worksheet.cell(row=start_row, column=1)
                            top_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        
                        start_row = excel_row
                        current_team = cell_value
            
            # Second: Merge money columns by filename (column M = 13)
            if len(df) > 1:
                start_row = 2  # Excel row 2 (after header)
                current_file = worksheet.cell(row=2, column=13).value
                
                for excel_row in range(3, worksheet.max_row + 2):  # +2 to include last row check
                    if excel_row > worksheet.max_row:
                        cell_value = None
                    else:
                        cell_value = worksheet.cell(row=excel_row, column=13).value
                    
                    if cell_value != current_file:
                        # End of group - merge if group size > 1
                        end_row = excel_row - 1
                        if end_row > start_row:
                            for col_idx in merge_by_file_cols:
                                col_letter = get_column_letter(col_idx)
                                worksheet.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
                                # Set alignment for merged cell
                                top_cell = worksheet.cell(row=start_row, column=col_idx)
                                top_cell.alignment = Alignment(horizontal="right", vertical="center")
                        
                        start_row = excel_row
                        current_file = cell_value

        output.seek(0)
        
        st.download_button(
            label="💾 Tải file Excel kết quả",
            data=output,
            file_name="hoadon_tonghop.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True
        )

        st.divider()
        st.dataframe(df, use_container_width=True)

    else:
        # === STEP 1: REQUIRED INPUTS ===
        st.markdown("### 📝 Bước 1: Thông tin bắt buộc")
        
        col1, col2 = st.columns(2)
        with col1:
            team_input = st.text_input("Team *", placeholder="Ví dụ: Team A, Team B...")
        with col2:
            employee_input = st.text_input("Tên nhân viên *", placeholder="Ví dụ: Nguyễn Văn A...")
        
        # === STEP 2: OPTIONAL CLASSIFICATION ===
        st.markdown("### 🏷️ Bước 2: Phân loại (Tùy chọn)")
        
        col_cat1, col_cat2 = st.columns(2)
        with col_cat1:
            category_select = st.selectbox("Chọn phân loại:", CATEGORY_OPTIONS)
        with col_cat2:
            custom_category = ""
            if category_select == "Khác (Nhập tay)":
                custom_category = st.text_input("Nhập phân loại tùy chỉnh:")
        
        st.divider()
        
        # === STEP 3: FILE UPLOAD ===
        st.markdown("### 📂 Bước 3: Tải hóa đơn (PDF)")
        
        # Check if required inputs are filled
        can_upload = bool(team_input.strip()) and bool(employee_input.strip())
        
        if not can_upload:
            st.warning("⚠️ Vui lòng nhập **Team** và **Tên nhân viên** trước khi tải file!")
        
        uploaded_files = st.file_uploader(
            "Kéo thả hoặc chọn nhiều file PDF vào đây", 
            type="pdf", 
            accept_multiple_files=True,
            disabled=not can_upload
        )

        if uploaded_files:
            st.divider()
            st.markdown("### ⚙️ Bước 4: Xử lý dữ liệu")
            st.write(f"Đã chọn **{len(uploaded_files)}** file.")
            
            if st.button("🚀 Bắt đầu trích xuất dữ liệu", type="primary"):
                logger.info(f"--- ACTION: Team={team_input}, Employee={employee_input} started processing {len(uploaded_files)} files ---")
                
                progress_bar = st.progress(0)
                status_box = st.empty()
                
                all_rows = []
                
                for i, uploaded_file in enumerate(uploaded_files):
                    status_box.info(f"⏳ Đang xử lý: **{uploaded_file.name}** ({i+1}/{len(uploaded_files)})")
                    progress_bar.progress((i + 1) / len(uploaded_files))
                    
                    try:
                        data, line_items = extract_invoice_data(uploaded_file, filename=uploaded_file.name)
                        uploaded_file.seek(0)
                        
                        # Determine classification
                        if category_select == "Khác (Nhập tay)" and custom_category.strip():
                            final_category = custom_category.strip()
                        elif category_select == "Tự động nhận diện":
                            # Auto-detect based on invoice content
                            if line_items:
                                all_item_names = " ".join([item.get("name", "") for item in line_items])
                                final_category = classify_content(all_item_names, data.get("Đơn vị bán", ""))
                            else:
                                final_category = classify_content("", data.get("Đơn vị bán", ""))
                        else:
                            final_category = category_select
                        
                        # Determine tax rate(s)
                        tax_rates = []
                        for rate in ["0%", "5%", "8%", "10%"]:
                            col_name = f"Thuế {rate}"
                            if data.get(col_name) and data.get(col_name) != "":
                                tax_rates.append(rate)
                        
                        if data.get("Thuế khác"):
                            tax_rates.append("Khác")
                        
                        if not tax_rates:
                            tax_rates = ["N/A"]
                        
                        # Create row(s) for this invoice
                        base_row = {
                            "Team": team_input.strip(),
                            "Số hóa đơn": data.get("Số hóa đơn", ""),
                            "Ngày hóa đơn": data.get("Ngày hóa đơn", ""),
                            "Mã số thuế bên bán": data.get("Mã số thuế", ""),
                            "Số ký hiệu": data.get("Ký hiệu", ""),
                            "Link tra cứu": data.get("Link lấy hóa đơn", "") or data.get("Mã tra cứu", ""),
                            "Phân loại": final_category,
                            "Số tiền trước VAT": data.get("Số tiền trước Thuế", ""),
                            "Tổng tiền sau thuế": data.get("Số tiền sau", ""),
                            "Tên nhân viên": employee_input.strip(),
                            "Tên file": uploaded_file.name
                        }
                        
                        # Handle multi-rate invoices
                        if len(tax_rates) == 1:
                            # Single rate - simple case
                            rate = tax_rates[0]
                            if rate == "N/A":
                                base_row["VAT"] = data.get("Tiền thuế", "")
                                base_row["Thuế suất"] = ""
                            else:
                                base_row["VAT"] = data.get(f"Thuế {rate}", data.get("Tiền thuế", ""))
                                base_row["Thuế suất"] = rate
                            all_rows.append(base_row)
                        else:
                            # Multiple rates - create multiple rows
                            for rate in tax_rates:
                                row = base_row.copy()
                                if rate == "Khác":
                                    row["VAT"] = data.get("Thuế khác", "")
                                    row["Thuế suất"] = "Khác"
                                else:
                                    row["VAT"] = data.get(f"Thuế {rate}", "")
                                    row["Thuế suất"] = rate
                                all_rows.append(row)
                        
                    except Exception as e:
                        logger.error(f"Error processing {uploaded_file.name}: {e}")
                        status_box.error(f"Lỗi khi xử lý {uploaded_file.name}")
                
                status_box.success("✅ Đã xử lý xong tất cả!")
                logger.info(f"--- COMPLETION: User {current_user} finished processing ---")
                
                # Create DataFrame with new column order
                columns = [
                    "Team", "Số hóa đơn", "Ngày hóa đơn", "Mã số thuế bên bán", 
                    "Số ký hiệu", "Link tra cứu", "Phân loại", 
                    "Số tiền trước VAT", "VAT", "Thuế suất", "Tổng tiền sau thuế",
                    "Tên nhân viên", "Tên file"
                ]
                df = pd.DataFrame(all_rows)
                for col in columns:
                    if col not in df.columns:
                        df[col] = ""
                df = df[columns]
                
                # Convert money columns
                money_columns = ["Số tiền trước VAT", "VAT", "Tổng tiền sau thuế"]
                for col in money_columns:
                    def convert_to_number(x):
                        if pd.isna(x) or x == '': return None
                        x_str = str(x).strip()
                        if re.search(r',\d{2}$', x_str):
                            x_str = x_str.replace('.', '').replace(',', '.')
                        else:
                            x_str = x_str.replace('.', '').replace(',', '')
                        try:
                            return round(float(x_str))
                        except:
                            return x
                    df[col] = df[col].apply(convert_to_number)
                
                # Save to session state
                st.session_state["processed_df"] = df
                st.session_state["processing_complete"] = True
                st.rerun()
        else:
            st.info("👆 Vui lòng tải file lên để tiếp tục.")
